import pandas as pd
from openpyxl import load_workbook
import pyodbc


# Cadena de conexión
connection_string = "Driver={SQL Server};" \
                    "Server=192.168.1.00;" \
                    "Database=DB_TEST;" \
                    "UID=USER;" \
                    "PWD=PASS"

query = "SELECT CODIGO, PATENTE, ID, CHASIS, COMPLEMENTOS FROM MOVILES WITH(NOLOCK) WHERE ID <> '9999' AND CODIGO NOT LIKE '%-%'"

#============================================================
ruta_archivo = r'C:\Users\andres\zrac310723.xlsx'
dataframe = pd.read_excel(ruta_archivo)

nuevo_ruta_archivo = r'C:\Users\andres\ZRAC_FLOTA_31072023_modified.xlsx'

# Eliminar columnas restantes
columnas_deseadas = ["Patente", "Serie"]
dataframe = dataframe[columnas_deseadas]

dataframe.insert(dataframe.columns.get_loc("Serie"), "VIN", "")
dataframe.insert(dataframe.columns.get_loc("Serie"), "PAT", "")
dataframe.insert(dataframe.columns.get_loc("Serie"), "SINF", "")
dataframe.insert(dataframe.columns.get_loc("Serie"), "SINR", "")

dataframe.to_excel(nuevo_ruta_archivo, index=False)
workbook = load_workbook(nuevo_ruta_archivo)
sheet = workbook.active

# Obtener el número de filas de las columnas
filas = sheet.max_row

# Agregar las fórmulas
for fila in range(2, filas + 1):
    celda_serie = f"F{fila}"
    celda_sinr = f"E{fila}"
    formula = f'=IF(OR(RIGHT({celda_serie}, 1)="r", RIGHT({celda_serie}, 1)="R"), LEFT({celda_serie}, LEN({celda_serie})-1), {celda_serie})'
    sheet[celda_sinr].value = formula

for fila in range(2, filas + 1):
    celda_sinr = f"E{fila}"
    celda_sinf = f"D{fila}"
    formula_sinf = f'=IF(OR(RIGHT({celda_sinr}, 1)="f", RIGHT({celda_sinr}, 1)="F"), LEFT({celda_sinr}, LEN({celda_sinr})-1), {celda_sinr})'
    sheet[celda_sinf].value = formula_sinf

for fila in range(2, filas + 1):
    celda_patente = f"A{fila}"
    celda_pat = f"C{fila}"
    formula_pat = f'=CONCATENATE(LEFT({celda_patente},4),"-",RIGHT({celda_patente},2))'
    sheet[celda_pat].value = formula_pat

for fila in range(2, filas + 1):
    celda_sinf = f"D{fila}"
    celda_vin = f"B{fila}"
    formula_vin = f'=MID({celda_sinf},11,200)'
    sheet[celda_vin].value = formula_vin

workbook.save(nuevo_ruta_archivo)
# Ejecutar la consulta
conn = pyodbc.connect(connection_string)

try:
    dataframe = pd.read_sql_query(query, conn)
    data = dataframe.values
    
    # Obtener los encabezados de las columnas
    headers = dataframe.columns
    
    # Pegar los encabezados en el archivo de Excel
    for j, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=8+j).value = header

    for i, row in enumerate(data, start=2):
        for j, value in enumerate(row, start=2):
            sheet.cell(row=i, column=7+j).value = value

    print("La consulta se ejecutó correctamente.")

except Exception as e:
    print("Error al ejecutar la consulta:", e)

conn.close()

sheet.insert_cols(10)
sheet.cell(row=1, column=10).value = "MATCH"
for i in range(2, len(data) + 2):
        sheet.cell(row=i, column=10).value = f'=VLOOKUP(I{i},$B$2:$C$21930,2,FALSE)'
        

sheet.insert_cols(15)
sheet.cell(row=1, column=15).value = "SP"
for i in range(2, len(data) + 2):
    actual_movcodigo = sheet.cell(row=i, column=9).value
    nuevo_movcodigo = sheet.cell(row=i, column=10).value
    sheet.cell(row=i, column=15).value = f'="Exec [BD_TEST].[dbo].[EncolarNuevoCambio_Sp]@Actual_Codigo=\'{actual_movcodigo}\',@Nuevo_Codigo=\'" & J{i} & "\'"'

# Guardar y cerrar el archivo de Excel
workbook.save(nuevo_ruta_archivo)
workbook.close()